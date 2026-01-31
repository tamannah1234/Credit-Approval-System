# core/tasks.py
from celery import shared_task
import openpyxl
from django.utils import timezone
from .models import Customer, Loan


@shared_task
def ingest_customer_data(file_path):
    """
    Ingests customer data from an Excel file into the database.
    Expected columns (in order):
    customer_id, first_name, last_name, age, phone_number, monthly_salary, approved_limit
    current_debt defaults to 0 if missing.
    Rows missing required fields (first_name, last_name, age) will be skipped.
    """
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Ensure row has enough columns
        if len(row) < 7:
            continue  # skip incomplete rows

        customer_id, first_name, last_name, age, phone_number, monthly_salary, approved_limit, *rest = row

        # Skip rows with missing required fields
        if not first_name or not last_name or age is None:
            continue

        current_debt = rest[0] if rest else 0  # default to 0 if missing
        if current_debt is None:
            current_debt = 0

        defaults = {
            'first_name': first_name,
            'last_name': last_name,
            'age': age,
            'phone_number': str(phone_number) if phone_number else None,
            'monthly_salary': monthly_salary or 0,
            'approved_limit': approved_limit or 0,
            'current_debt': current_debt,
        }

        if customer_id:
            Customer.objects.update_or_create(id=customer_id, defaults=defaults)
        else:
            Customer.objects.create(**defaults)
            
@shared_task(bind=True, autoretry_for=(Exception,), retry_backoff=5, retry_kwargs={"max_retries": 3})
def ingest_loan_data(self, file_path):
    """
    Ingests loan data from an Excel file into the database.

    Expected columns (in order):
    customer_id, loan_id, loan_amount, tenure,
    interest_rate, emi, emis_paid, start_date, end_date
    """
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2, values_only=True):

        # Skip invalid / incomplete rows
        if not row or len(row) != 9:
            continue

        (
            customer_id,
            loan_id,
            loan_amount,
            tenure,
            interest_rate,
            emi,
            emis_paid,
            start_date,
            end_date
        ) = row

        # Skip if customer does not exist
        try:
            customer = Customer.objects.get(id=int(customer_id))
        except (Customer.DoesNotExist, TypeError, ValueError):
            continue

        Loan.objects.update_or_create(
            id=int(loan_id),
            defaults={
                "customer": customer,
                "loan_amount": float(loan_amount),
                "tenure": int(tenure),
                "interest_rate": float(interest_rate),
                "monthly_repayment": float(emi),
                "emis_paid_on_time": int(emis_paid) if emis_paid else 0,
                "start_date": start_date or timezone.now(),
                "end_date": end_date,
            }
        )
